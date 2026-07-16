#!/usr/bin/env python3
"""
Audit <STUDY>/<ID>/<SUBFOLDER> trees on an rclone remote (ownCloud / WebDAV).

Reads directory metadata only - file contents are never downloaded, so this is
safe to run against privatecloud without staging anything on the cluster.

Usage:
    python3 wgs_check.py privatecloud:ENIGMA -o enigma_wgs.csv --xlsx enigma_wgs.xlsx
    python3 wgs_check.py privatecloud:OTHER_STUDY -o other.csv
    python3 wgs_check.py privatecloud:STUDY --subfolder WES

Every ID directly under the study folder gets at least one row, including IDs
whose WGS folder is missing or empty - those are the ones you want to find.
"""

import argparse
import csv
import fnmatch
import json
import os
import re
import subprocess
import sys

GB = 1024 ** 3

FASTQ_SUFFIXES = (".fastq.gz", ".fq.gz")

# Trailing R1/R2 marker: _R1, .R2, -1, _R1_001 ... The separator prefix keeps
# digits inside the sample name itself (P001, L002) from matching.
READ_RE = re.compile(r"[._-][Rr]?([12])(?:[._-]|$)")

# Naming varies by study: checksum.txt, 2545_checksums.txt, md5sums.txt ...
CHECKSUM_GLOB = "*checksum*.txt"

FIELDNAMES = [
    "sample_id",
    "file_name",
    "size_gb",
    "read",
    "has_checksum",
    "checksum_file",
    "fastq_gz_count",
    "r1_count",
    "r2_count",
    "size_gb_diff",
    "complete",
    "note",
]


def list_remote(root, rclone):
    """One recursive listing of the study tree: directories and files."""
    cmd = [rclone, "lsjson", root, "--recursive"]
    try:
        proc = subprocess.run(cmd, check=True, capture_output=True, text=True)
    except FileNotFoundError:
        sys.exit(f"error: rclone not found at '{rclone}'.\n"
                 "Pass the full path with --rclone (on the cluster it is "
                 "/usr/bin/rclone).")
    except subprocess.CalledProcessError as e:
        sys.exit(f"error: rclone failed listing {root}:\n{e.stderr.strip()}")
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError as e:
        sys.exit(f"error: could not parse rclone output: {e}")


def is_fastq_gz(name):
    return name.lower().endswith(FASTQ_SUFFIXES)


def read_number(name):
    """1 or 2 for an R1/R2 file, or None if the name doesn't say."""
    stem = name
    for suf in FASTQ_SUFFIXES:
        if stem.lower().endswith(suf):
            stem = stem[:-len(suf)]
            break
    matches = READ_RE.findall(stem)
    return int(matches[-1]) if matches else None


def collect(entries, subfolder):
    """Group the flat listing into one record per ID."""
    sub_lc = subfolder.lower()
    ids = {}

    # Pass 1: every directory directly under the study root is an ID.
    for e in entries:
        parts = e["Path"].split("/")
        if e["IsDir"] and len(parts) == 1:
            ids[parts[0]] = {"has_subfolder": False, "files": []}

    # Pass 2: locate the <ID>/<SUBFOLDER> directory and its immediate files.
    for e in entries:
        parts = e["Path"].split("/")
        if len(parts) < 2 or parts[0] not in ids:
            continue
        rec = ids[parts[0]]
        if e["IsDir"]:
            if len(parts) == 2 and parts[1].lower() == sub_lc:
                rec["has_subfolder"] = True
        elif len(parts) == 3 and parts[1].lower() == sub_lc:
            rec["files"].append(e)

    return ids


def build_rows(ids, subfolder, checksum_glob):
    glob_lc = checksum_glob.lower()
    rows = []

    for sample_id in sorted(ids):
        rec = ids[sample_id]
        files = rec["files"]
        fastqs = sorted((f for f in files if is_fastq_gz(f["Name"])),
                        key=lambda f: f["Name"])
        checksums = sorted(f["Name"] for f in files
                           if fnmatch.fnmatch(f["Name"].lower(), glob_lc))
        has_checksum = bool(checksums)
        count = len(fastqs)

        reads = {f["Name"]: read_number(f["Name"]) for f in fastqs}
        r1 = sum(1 for v in reads.values() if v == 1)
        r2 = sum(1 for v in reads.values() if v == 2)
        unknown = sum(1 for v in reads.values() if v is None)

        # Reads must pair up: N x R1 against N x R2. Two lanes (4 files) is a
        # healthy sample, so completeness turns on the balance, not the total.
        paired = r1 >= 1 and r1 == r2 and unknown == 0
        complete = "yes" if (has_checksum and paired) else "no"

        # R1 and R2 of a real pair hold the same number of reads, so their
        # sizes track closely. A large gap means one side is truncated.
        # Computed from raw bytes, not the rounded size_gb column.
        if paired:
            r1_bytes = sum(f["Size"] for f in fastqs if reads[f["Name"]] == 1)
            r2_bytes = sum(f["Size"] for f in fastqs if reads[f["Name"]] == 2)
            size_diff = round(abs(r1_bytes - r2_bytes) / GB, 2)
        else:
            size_diff = ""

        if not rec["has_subfolder"]:
            note = f"no {subfolder} folder"
        elif count == 0:
            note = f"{subfolder} folder is empty" if not files else "no fastq.gz files"
        elif unknown:
            note = f"cannot tell R1/R2 from {unknown} file name(s)"
        elif r1 != r2:
            note = f"unbalanced: {r1} x R1 vs {r2} x R2"
        elif not has_checksum:
            note = f"no file matching {checksum_glob}"
        elif count > 2:
            note = f"{r1} R1/R2 pairs (multi-lane) - counts balance"
        else:
            note = ""

        base = {
            "sample_id": sample_id,
            "has_checksum": "TRUE" if has_checksum else "FALSE",
            "checksum_file": " | ".join(checksums),
            "fastq_gz_count": count,
            "r1_count": r1,
            "r2_count": r2,
            "size_gb_diff": size_diff,
            "complete": complete,
            "note": note,
        }

        if fastqs:
            for f in fastqs:
                rows.append({**base,
                             "file_name": f["Name"],
                             "size_gb": round(f["Size"] / GB, 2),
                             "read": reads[f["Name"]] or "?"})
        else:
            # Nothing to list, but the ID must still be visible in the report.
            rows.append({**base, "file_name": "", "size_gb": "", "read": ""})

    return rows


def write_csv(rows, path):
    with open(path, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("error: --xlsx needs openpyxl. Install it with:\n"
                 "    pip install openpyxl")

    band_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="4472C4")

    wb = Workbook()
    ws = wb.active
    ws.title = "WGS audit"
    ws.append(FIELDNAMES)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Shade alternate samples rather than alternate rows, so the R1/R2 lines of
    # one sample read as a single block.
    banded = False
    prev_id = None
    for r in rows:
        if r["sample_id"] != prev_id:
            banded = not banded
            prev_id = r["sample_id"]
        ws.append([r[k] for k in FIELDNAMES])
        if banded:
            for cell in ws[ws.max_row]:
                cell.fill = band_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{len(rows) + 1}"
    for i, name in enumerate(FIELDNAMES, start=1):
        longest = max([len(name)] + [len(str(r[name])) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(longest + 2, 60)

    wb.save(path)


def main():
    ap = argparse.ArgumentParser(
        description="Audit <STUDY>/<ID>/<SUBFOLDER> on an rclone remote.")
    ap.add_argument("root", help="study folder, e.g. privatecloud:ENIGMA")
    ap.add_argument("-o", "--out", default="wgs_report.csv",
                    help="output CSV (default: wgs_report.csv)")
    ap.add_argument("--xlsx", metavar="FILE",
                    help="also write an Excel copy")
    ap.add_argument("--subfolder", default="WGS",
                    help="per-ID data folder (default: WGS)")
    ap.add_argument("--checksum-glob", default=CHECKSUM_GLOB,
                    help=f"checksum filename pattern, case-insensitive "
                         f"(default: {CHECKSUM_GLOB})")
    ap.add_argument("--rclone", default=os.environ.get("RCLONE", "rclone"),
                    help="rclone binary (default: $RCLONE, else 'rclone')")
    args = ap.parse_args()

    entries = list_remote(args.root, args.rclone)
    ids = collect(entries, args.subfolder)

    if not ids:
        sys.exit(f"error: no ID folders found under {args.root}.\n"
                 "Check the remote name and that the path points at the study "
                 "folder itself (the one that contains the ID folders).")

    rows = build_rows(ids, args.subfolder, args.checksum_glob)
    write_csv(rows, args.out)
    if args.xlsx:
        write_xlsx(rows, args.xlsx)

    incomplete = sorted({r["sample_id"] for r in rows if r["complete"] == "no"})
    print(f"IDs found:      {len(ids)}")
    print(f"Complete:       {len(ids) - len(incomplete)}")
    print(f"Incomplete:     {len(incomplete)}")
    print(f"\nWrote {args.out}" + (f" and {args.xlsx}" if args.xlsx else ""))

    if incomplete:
        print("\nIncomplete IDs:")
        seen = set()
        for r in rows:
            sid = r["sample_id"]
            if r["complete"] == "no" and sid not in seen:
                seen.add(sid)
                print(f"  {sid:<20} {r['note']}")


if __name__ == "__main__":
    main()
